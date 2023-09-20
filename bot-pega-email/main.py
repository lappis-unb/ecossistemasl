import pandas as pd
import openpyxl
import os
import platform
import time
from selenium import webdriver
from selenium.webdriver.common.by import By

xlsx_file = 'PPA Participativo Tratamento de Demandas.xlsx'
driver = webdriver.Chrome()
somar_coluna = False
time_sleep = 2


def realiza_login():
    url = 'https://sso.acesso.gov.br/login?client_id=brasilparticipativo.presidencia.gov.br&authorization_id=18aa998b34f#'
    driver.get(url)
    input("Faca o login e de ENTER...")


def busca_email(proposta, linha, coluna, votos):
    try:
        driver.get('https://brasilparticipativo.presidencia.gov.br/admin/participatory_processes/programas/components/2'
                   '/manage/')

        xpath_busca_prop = '//*[@id="q_id_string_or_title_cont"]'
        element = driver.find_element(By.XPATH, xpath_busca_prop)
        element.send_keys(proposta)

        xpath_buscar = '//*[@id="decidim/proposals/proposal_search"]/div/div/button'
        element = driver.find_element(By.XPATH, xpath_buscar)
        element.click()

        selector_lista_prop = '[data-id]'
        response_id = verificar_tabela(driver, selector_lista_prop, votos)

        if response_id[1] == -1:
            print(response_id[0])
            adicionar_id(-1)
            return

        url_prop = (
            f'https://brasilparticipativo.presidencia.gov.br/admin/participatory_processes/programas/components/2'
            f'/manage/proposals/{response_id[1]}')

        driver.get(url_prop)

        driver.get(verifica_autor(driver))

        time.sleep(time_sleep)

        xpath_abre_email = '//*[@id="user-groups"]/div[4]/div/table/tbody/tr/td[7]/a[2]/span'
        element = driver.find_element(By.XPATH, xpath_abre_email)
        element.click()

        time.sleep(time_sleep)

        xpath_mostra_email = '//*[@id="show-email-modal"]/div[2]/div/div[2]/button'
        element = driver.find_element(By.XPATH, xpath_mostra_email)
        element.click()

        time.sleep(time_sleep)

        xpath_email = '//*[@id="user_email"]/a'
        element = driver.find_element(By.XPATH, xpath_email)
        response_email = element.text

        print(f'{linha}: {response_email}')

        atualiza_email(response_email, linha, coluna)

        adicionar_id(response_id[1])
    except:
        input('Error...')


def verifica_autor(driver):
    primeiro_elemento_a = driver.find_element(By.CSS_SELECTOR, '#proposal-authors-list a')
    href = primeiro_elemento_a.get_attribute("href")
    nome_usuario = href.split("/")[-1]
    return (f'https://brasilparticipativo.presidencia.gov.br/admin/officializations?utf8=%E2%9C%93&q'
            f'%5Bname_or_nickname_or_email_or_identities_uid_cont%5D={nome_usuario}')


def verificar_tabela(driver, selector_tabela, votos):
    tabela = driver.find_elements(By.CSS_SELECTOR, selector_tabela)
    votos = int(votos)

    if len(tabela) == 0:
        return ["A tabela está vazia.", -1]
    elif len(tabela) == 1:
        elemento = tabela[0]
        #print(elemento.find_element(By.XPATH, './td[6]').text)
        # Verifica se o número de votos é igual ao valor desejado
        if int(elemento.find_element(By.XPATH, './td[6]').text) == votos:
            return ["", elemento.get_attribute("data-id")]
        else:
            return [f"A única linha na tabela não possui votos iguais a {votos}.", -1]
    else:
        for elemento in tabela:
            #print(elemento.find_element(By.XPATH, './td[6]').text)
            # Verifica se o número de votos é igual ao valor desejado
            if int(elemento.find_element(By.XPATH, './td[6]').text) == votos:
                return ["", elemento.get_attribute("data-id")]
        return [f"Nenhuma linha na tabela possui votos iguais a {votos}.", -1]


def adicionar_id(response_id):
    global somar_coluna
    if sheet.cell(row=1, column=1).value.startswith('id'):
        sheet.cell(row=linha, column=1, value=response_id)
    else:
        sheet.insert_cols(1)
        sheet.cell(row=1, column=1, value='id')
        sheet.cell(row=linha, column=1, value=response_id)

        if somar_coluna == False:
            somar_coluna = True


def atualiza_email(response_email, linha, coluna):
    sheet.cell(row=linha, column=coluna, value=response_email)
    workbook.save(xlsx_file)


def limpar_terminal():
    sistema = platform.system()
    if sistema == "Windows":
        os.system('cls')
    else:
        os.system('clear')


workbook = openpyxl.load_workbook(xlsx_file)
sheet_names = workbook.sheetnames

limpar_terminal()
print("Planilhas disponíveis:")
for i, sheet_name in enumerate(sheet_names):
    print(f"{i + 1}: {sheet_name}")

selected_sheet_index = int(input("Digite o número da planilha que deseja usar: ")) - 1
selected_sheet_name = sheet_names[selected_sheet_index]

sheet = workbook[selected_sheet_name]

header_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]

limpar_terminal()
print("Colunas disponíveis:")
for i, column_name in enumerate(header_row):
    print(f"{i + 1}: {column_name}")

selected_data_column_index = int(input("Digite o número da coluna da qual deseja retirar os dados: ")) - 1
selected_print_column_index = int(input("Digite o número da coluna que deseja imprimir: ")) - 1
selected_votos_column_index = int(input("Digite o numero da coluna que possui os votos: ")) - 1

linha = 2

realiza_login()
somado = False
for row in sheet.iter_rows(min_row=2, values_only=True):
    if somar_coluna == True and somado == False:
        selected_data_column_index += 1
        selected_votos_column_index += 1
        selected_print_column_index += 1
        somado = True

    data_value = row[selected_data_column_index]
    votos_value = row[selected_votos_column_index]
    busca_email(data_value, linha, selected_print_column_index + 1, votos_value)
    linha += 1
    time.sleep(time_sleep)

workbook.close()
driver.quit()
