from tkinter import filedialog
from openpyxl.styles import PatternFill, Font
from openpyxl.styles.borders import Border, Side

import tkinter as tk
import pyperclip
import openpyxl
import os
import json
import traceback


def organize_xlsx():
    file = file_entry.get()
    if file:
        try:
            sheet_name = create_sheet(file)
            response_label.config(text=f"Arquivo organizado com sucesso em: '{sheet_name}'.")
        except Exception as e:
            error_line = traceback.extract_tb(e.__traceback__)[-1].lineno
            response_label.config(text=f"Erro na linha {error_line}: {str(e)}")
    else:
        response_label.config(text="Por favor, escolha um arquivo 'XLSX' primeiro.")


def create_sheet(file):
    try:
        # Open the original Excel file
        workbook = openpyxl.load_workbook(file)
        original_sheet = workbook.active

        # Initialize the count of non-empty rows
        num_rows = 0

        # Find the number of non-empty rows
        for row in original_sheet.iter_rows():
            if any(cell.value for cell in row):
                num_rows += 1

        # Create a new XLSX file with the original name + "_FORMATTED"
        formatted_name = file.replace('.xlsx', '_FORMATADO.xlsx')
        unique_name = get_unique_name(formatted_name)
        new_workbook = openpyxl.Workbook()

        # Remove empty sheet
        new_workbook.remove(new_workbook.active)

        # Create numbered sheets from 1 to the number of rows in the new file
        for number in range(1, num_rows):
            new_sheet = new_workbook.create_sheet(title=str(number))

            # Copy the first row
            first_row = [cell.value for cell in original_sheet[1]]

            # Copy the row at position "number"
            number_row = [cell.value for cell in original_sheet[number + 1]]

            # Add the rows to the new sheet
            new_sheet.append(first_row)
            new_sheet.append(number_row)

            unpack_proposal(new_sheet)

            fill = PatternFill(start_color="c0c0c0", end_color="c0c0c0", fill_type="solid")
            bold_font = Font(bold=True)
            border_style = Border(
                left=Side(border_style="thin"),
                right=Side(border_style="thin"),
                top=Side(border_style="thin"),
                bottom=Side(border_style="thin")
            )
            alignment = openpyxl.styles.Alignment(wrap_text=True, horizontal='center', vertical='center')

            for i, row in enumerate(new_sheet.iter_rows(min_row=1, max_row=new_sheet.max_row)):
                if i == 0:
                    for cell in row:
                        cell.fill = fill
                        cell.font = bold_font
                        cell.border = border_style
                        cell.alignment = alignment
                else:
                    for cell in row:
                        if cell.value is not None:
                            cell.border = border_style
                            cell.alignment = alignment

            for i in range(0, new_sheet.max_column):
                new_sheet.column_dimensions[decimal_to_alphabet(i)].width = 25

            # Save the new file
            new_workbook.save(unique_name)

        return unique_name

    except Exception as e:
        raise e


def unpack_proposal(sheet):
    column_index = 32
    column_position = decimal_to_alphabet(column_index)
    row_position = 1

    column = sheet[column_position]
    cell = column[row_position]

    if cell is None or cell.value is None:
        return

    data = json.loads(cell.value)

    # Create 5 columns to the right of 'AB'
    for i in range(5):
        sheet.insert_cols(column[-1].column + 1)

    # Define the titles to be written with the JSON keys mappings
    titles = [
        {"Proposal Title": "title"},
        {"Main Challenge to Overcome": "identify_challenge"},
        {"Main Action to Take": "describe"},
        {"Main Federal Public Entity to Lead the Action": "identify_entity"},
        {"Thematic Axis": "axle"}
    ]

    for i, title in enumerate(titles):
        title_cell = sheet.cell(row=1, column=column_index + 2 + i)
        title_cell.value = list(title.keys())[0]

        current_row = 2
        for item in data:
            key = list(title.values())[0]
            value = item.get(key, "")
            if value:
                cell = sheet.cell(row=current_row, column=column_index + 2 + i)
                cell.value = value
                enumerator = sheet.cell(row=current_row, column=column_index + 1)
                enumerator.value = current_row - 1
                current_row += 1


def choose_file():
    file = filedialog.askopenfilename(filetypes=[("XLSX Files", "*.xlsx")])
    file_entry.delete(0, tk.END)
    file_entry.insert(0, file)


def get_unique_name(file):
    count = 1
    base_name, extension = os.path.splitext(file)

    while os.path.exists(file):
        file = f"{base_name}_{count}{extension}"
        count += 1

    return file


def decimal_to_alphabet(decimal):
    alphabet = ""
    while decimal >= 0:
        remainder = decimal % 26
        alphabet = chr(65 + remainder) + alphabet
        decimal = (decimal - remainder) // 26
        if decimal == 0:
            break
        else:
            decimal -= 1
    return alphabet


def complete_links():
    file = file_entry.get()
    if file:
        try:
            workbook = openpyxl.load_workbook(file)

            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                update_cells_with_text(worksheet,
                                       '/rails/active_storage/blobs',
                                       'https://brasilparticipativo.presidencia.gov.br')

            workbook.save(file)
            response_label.config(text=f"Links atualizados com sucesso.")
        except Exception as e:
            error_line = traceback.extract_tb(e.__traceback__)[-1].lineno
            response_label.config(text=f"Erro na linha {error_line}: {str(e)}")
    else:
        response_label.config(text="Por favor, escolha um arquivo 'XLSX' primeiro.")


def update_cells_with_text(sheet, search_text, prefix):
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value and search_text in str(cell.value) and prefix not in str(cell.value):
                cell.style = 'Hyperlink'
                cell.hyperlink = prefix + cell.value


def copy_output():
    output = response_label.cget("text")
    if output:
        pyperclip.copy(output)


# Main window setup
window = tk.Tk()
window.title("Organizador de planilha")

# Title
title_label = tk.Label(window, text="Escolha um arquivo com formato 'XLSX'")
title_label.pack(pady=10)

# Button to choose a file
choose_file_button = tk.Button(window, text="Escolher arquivo XLSX", command=choose_file)
choose_file_button.pack()

# Entry field to display the chosen file name
file_entry = tk.Entry(window)
file_entry.pack(pady=10)

# Button to organize the XLSX file
organize_button = tk.Button(window, text="Organizar planilha", command=organize_xlsx)
organize_button.pack()

# Button to complete links the XLSX file
organize_button = tk.Button(window, text="Completar links", command=complete_links)
organize_button.pack()

# Label to display system response
response_label = tk.Label(window, text="")
response_label.pack(fill="both", expand=True, pady=10)

# Button to copy the output
copy_output_button = tk.Button(window, text="Copiar saida", command=copy_output)
copy_output_button.pack()

# Main GUI loop
window.mainloop()
